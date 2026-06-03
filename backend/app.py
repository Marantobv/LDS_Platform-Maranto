from flask import Flask, request, jsonify, send_file
import os
from flask_cors import CORS
import openpyxl
from openpyxl.styles import PatternFill, Font
import io
import re
import requests
import unicodedata
from datetime import date

app = Flask(__name__)
CORS(app)

RAILWAY_API = "https://web-production-ae47c.up.railway.app/dni"
RAILWAY_SECRET_KEY = "LDS_DNI"
# ── Helpers ───────────────────────────────────────────────────────────────────

def is_red_font(cell):
    try:
        font = cell.font
        if font is None:
            return False
        color = font.colors
        if color is None:
            return False
        if color.type == "rgb" and color.rgb:
            rgb = color.rgb.upper()
            if rgb.startswith("FF") and len(rgb) == 8:
                r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
                if r > 180 and g < 100 and b < 100:
                    return True
            if len(rgb) == 6:
                r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
                if r > 180 and g < 100 and b < 100:
                    return True
        if color.type == "indexed" and color.indexed == 2:
            return True
    except Exception:
        pass
    return False


def cell_value(cell):
    val = cell.value
    if val is None:
        return ""
    return str(val).strip()


def normalize(text):
    """Quita tildes, espacios extra y pasa a mayúsculas para comparar."""
    text = text.strip().upper()
    text = " ".join(text.split())  # normaliza espacios múltiples
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def build_excel_format_name(resultado):
    """Construye el nombre en formato Excel: APELLIDO_PAT APELLIDO_MAT NOMBRES"""
    ap = resultado.get("apellido_paterno", "").strip()
    am = resultado.get("apellido_materno", "").strip()
    nombres = resultado.get("nombres", "").strip()
    return f"{ap} {am} {nombres}".strip()


def query_dni(dni):
    """Consulta la API de Railway para un DNI. Retorna (nombre_api, error)."""
    try:
        dni = str(dni).strip().zfill(8)  # Rellena con ceros hasta 8 dígitos
        res = requests.get(f"{RAILWAY_API}/{dni}",
                        headers={"Authorization": f"Bearer {RAILWAY_SECRET_KEY}"},
                        timeout=8)
        if res.status_code != 200:
            return None, f"HTTP {res.status_code}"
        data = res.json()
        if not data.get("estado"):
            return None, data.get("mensaje", "No encontrado")
        nombre = build_excel_format_name(data["resultado"])
        return nombre, None
    except requests.Timeout:
        return None, "Timeout"
    except Exception as e:
        return None, str(e)


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.route("/api/upload-excel", methods=["POST"])
def upload_excel():
    if "file" not in request.files:
        return jsonify({"error": "No se recibió ningún archivo"}), 400

    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Solo se aceptan archivos .xlsx"}), 400

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        rows_data = []
        header_row = None

        for row in ws.iter_rows():
            if header_row is None:
                non_empty = [c for c in row if c.value is not None]
                if non_empty:
                    seen_headers = {}
                    header_row = []
                    for c in row:
                        name = cell_value(c)
                        if name in seen_headers:
                            seen_headers[name] += 1
                            name = f"{name}_{seen_headers[name]}"
                        else:
                            seen_headers[name] = 1
                        header_row.append(name)
                continue

            values = [cell_value(c) for c in row]
            if all(v == "" for v in values):
                continue

            row_dict = {}
            red_cells = []
            seen = {}

            for col_idx, cell in enumerate(row):
                col_name = header_row[col_idx] if col_idx < len(header_row) else f"col_{col_idx}"
                if col_name in seen:
                    seen[col_name] += 1
                    col_name = f"{col_name}_{seen[col_name]}"
                else:
                    seen[col_name] = 1
                row_dict[col_name] = cell_value(cell)
                if is_red_font(cell):
                    red_cells.append(col_name)

            rows_data.append({
                "data": row_dict,
                "redCells": red_cells,
                "hasRed": len(red_cells) > 0,
            })

        if not rows_data:
            return jsonify({"error": "El archivo no contiene datos"}), 400

        return jsonify({
            "rows": rows_data,
            "total": len(rows_data),
            "columns": header_row,
        })

    except Exception as e:
        return jsonify({"error": f"Error al procesar el archivo: {str(e)}"}), 500


@app.route("/api/validar-dni", methods=["POST"])
def validar_dni():
    """
    Recibe una lista de filas con DNI y CLIENTE, consulta la API
    y devuelve la comparativa.

    Body JSON: { "rows": [{ "solicitud": "...", "dni": "...", "cliente": "..." }] }
    """
    body = request.get_json()
    if not body or "rows" not in body:
        return jsonify({"error": "Body inválido"}), 400

    results = []
    for item in body["rows"]:
        solicitud = item.get("solicitud", "")
        dni = str(item.get("dni", "")).strip().zfill(8)  # Asegurar 8 dígitos
        cliente_excel = item.get("cliente", "").strip()

        if not dni:
            results.append({
                "solicitud": solicitud,
                "dni": dni,
                "cliente_excel": cliente_excel,
                "nombre_api": "",
                "coincide": False,
                "error": "DNI vacío",
            })
            continue

        nombre_api, error = query_dni(dni)

        if error:
            results.append({
                "solicitud": solicitud,
                "dni": dni,
                "cliente_excel": cliente_excel,
                "nombre_api": "",
                "coincide": False,
                "error": error,
            })
            continue

        coincide = normalize(cliente_excel) == normalize(nombre_api)
        results.append({
            "solicitud": solicitud,
            "dni": dni,
            "cliente_excel": cliente_excel,
            "nombre_api": nombre_api,
            "coincide": coincide,
            "error": None,
        })

    return jsonify({"results": results})


@app.route("/api/descargar-comparativa", methods=["POST"])
def descargar_comparativa():
    """
    Recibe los resultados de la comparativa y genera un Excel descargable.
    Body JSON: { "results": [...] }
    """
    body = request.get_json()
    if not body or "results" not in body:
        return jsonify({"error": "Body inválido"}), 400

    results = body["results"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparativa DNI"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")

    headers = ["Solicitud", "DNI", "Nombre en Excel", "Nombre en API", "¿Coincide?", "Observación"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, item in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=item.get("solicitud", ""))
        ws.cell(row=row_idx, column=2, value=item.get("dni", ""))
        ws.cell(row=row_idx, column=3, value=item.get("cliente_excel", ""))
        ws.cell(row=row_idx, column=4, value=item.get("nombre_api", ""))

        if item.get("error"):
            ws.cell(row=row_idx, column=5, value="ERROR")
            ws.cell(row=row_idx, column=6, value=item["error"])
            fill = yellow_fill
        elif item.get("coincide"):
            ws.cell(row=row_idx, column=5, value="SÍ")
            ws.cell(row=row_idx, column=6, value="")
            fill = green_fill
        else:
            ws.cell(row=row_idx, column=5, value="NO")
            ws.cell(row=row_idx, column=6, value="Nombre no coincide")
            fill = red_fill

        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="comparativa_dni.xlsx",
    )

@app.route("/api/descargar-clasificacion", methods=["POST"])
def descargar_clasificacion():
    """
    Recibe las 3 secciones y genera un Excel con un bloque por sección.
    Body JSON: { "columns": [...], "sections": { "main": [...], "fise": [...], "nofise": [...] } }
    """
    body = request.get_json()
    if not body:
        return jsonify({"error": "Body inválido"}), 400

    columns = body.get("columns", [])
    sections = body.get("sections", {})

    import json
    for key in ["main", "fise", "nofise"]:
        rows = sections.get(key, [])
        if rows:
            print(f"\n=== {key} - primera fila ===")
            print(json.dumps(rows[0], ensure_ascii=False, indent=2))

    SECTION_LABELS = {
        "main": "SIN CLASIFICAR",
        "fise": "FISE",
        "nofise": "NO FISE",
    }

    header_font   = Font(bold=True, color="FFFFFF")
    header_fill   = PatternFill("solid", fgColor="2563EB")
    section_font  = Font(bold=True, color="FFFFFF")
    section_fills = {
        "main":    PatternFill("solid", fgColor="6B7280"),
        "fise":    PatternFill("solid", fgColor="16A34A"),
        "nofise":  PatternFill("solid", fgColor="2563EB"),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clasificación"

    current_row = 1

    for section_key in ["main", "fise", "nofise"]:
        rows = sections.get(section_key, [])
        if not rows:
            continue

        # Fila de título de sección
        label = SECTION_LABELS[section_key]
        title_cell = ws.cell(row=current_row, column=1, value=label)
        title_cell.font = section_font
        title_cell.fill = section_fills[section_key]
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(columns))
        current_row += 1

        # Fila de encabezados
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1

        # Filas de datos
        parent_fill = PatternFill("solid", fgColor="BBFAD0")
        red_font    = Font(bold=True, color="FF0000")
        parent_font = Font(bold=True)
        for row in rows:
            is_parent = row.get("isParent", False)
            red_cells = row.get("redCells", [])
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=col_idx,
                               value=row.get("data", {}).get(col_name, ""))
                if col_name in red_cells:
                    cell.font = red_font
                elif is_parent:
                    cell.fill = parent_fill
                    cell.font = parent_font
            current_row += 1

        # Fila vacía de separación
        current_row += 1

    # Ajustar ancho de columnas
    for col_idx, col in enumerate(ws.columns, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max((len(str(c.value or "")) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="clasificacion.xlsx",
    )


# ── PEM Helpers ───────────────────────────────────────────────────────────────

# Mapeo: columna base maestra → columna en avance PEM
PEM_COLUMN_MAP = {
    "EXPEDIENTE":                    "N° EXP.",
    "AGRUPACIONES":                  "AGRUPACIÓN DE VIVIENDA",
    "DISTRITO":                      "DISTRITO DE OBRA",
    "% DE AVANCE":                   "AVANCE",
    "FECHA PUSER (REDES BT)":        "Fecha P/S Real",
    "LOTES PLANO":                   "LOTES",
    "LOTES ELECTRIFICADOS":          "LOTES PROYECTO",
    "ESTADO DE OBRA":                "STATUS",
    "REPRESENTANTE POB":             "REPRESENTANTE",
    "TELEFONO":                      "TELÉFONO",
    "COGEPULP":                      "COGEPULP",
}

# Inverso: PEM header → base header
PEM_TO_BASE = {v: k for k, v in PEM_COLUMN_MAP.items()}

# Columnas almacenadas como decimal en la base maestra (formato porcentaje)
PERCENT_COLUMNS = {"% DE AVANCE"}

# Celda PEM vacía no sobrescribe un valor ya existente en la base
KEEP_IF_PEM_BLANK = {"% DE AVANCE", "FECHA PUSER (REDES BT)"}

# No sobrescribir en filas existentes (EXPEDIENTE ya encontrado en la base)
SKIP_ON_EXISTING_UPDATE = {"AGRUPACIONES"}

PUSER_CONFIRMATION_COLUMN = "CONFIRMACION DE PROYECTOS (PUSER)"

# PEM y base usan distintos términos para el mismo valor → no disparar actualización
COLUMN_VALUE_EQUIVALENTS = {
    "ESTADO DE OBRA": {
        "FINALIZADO": "TERMINADO",
        "TERMINADO": "TERMINADO",
        "PARALIZADO": "PARALIZADO",
        "PARALIZADO OBRA": "PARALIZADO",
        "PARALIZADO PROYECTO": "PARALIZADO",
    },
}

# Valor a escribir en la base cuando el PEM usa un sinónimo
COLUMN_WRITE_CANONICAL = {
    "ESTADO DE OBRA": {
        "TERMINADO": "Terminado",
        "PARALIZADO": "Paralizado",
    },
}


def _single_percent_to_decimal(n):
    """Convierte un número suelto a decimal Excel (75 → 0.75, 0.75 → 0.75)."""
    n = float(n)
    return round(n / 100, 6) if n > 1 else round(n, 6)


def parse_percent_to_decimal(val):
    """'75%' | '75' | 75 | 0.75  →  0.75 (decimal). '' | None → None."""
    if val is None or val == "":
        return None
    s = str(val).strip().replace("%", "").strip()
    if s == "":
        return None
    try:
        return _single_percent_to_decimal(float(s.replace(",", ".")))
    except ValueError:
        return None


def parse_pem_avance_to_decimal(val):
    """AVANCE PEM: '90%20%' (con saltos de línea) → menor %; un solo valor como la base."""
    if pem_cell_is_blank(val):
        return None
    s = normalize_header(str(val))
    matches = re.findall(r"(\d+(?:[.,]\d+)?)\s*%", s)
    if matches:
        return min(
            _single_percent_to_decimal(float(m.replace(",", ".")))
            for m in matches
        )
    return parse_percent_to_decimal(val)


def effective_avance_decimal(mapped_changes, base_row):
    """Avance efectivo para reglas (no usar un PEM menor si la base ya es mayor)."""
    new_dec = mapped_changes.get("% DE AVANCE")
    old_dec = parse_percent_to_decimal(base_row.get("% DE AVANCE"))
    if new_dec is None:
        return old_dec
    if old_dec is None:
        return new_dec
    return max(old_dec, new_dec)


def format_percent_display(val):
    """0.75 → '75%'  para mostrar en el preview"""
    if val is None or val == "":
        return "—"
    try:
        return f"{round(float(val) * 100, 2):g}%"
    except (ValueError, TypeError):
        return str(val)


def normalize_header(text):
    """Normaliza un header: quita saltos de línea, espacios extra, strip."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\n", " ").replace("\r", " ").split()).strip()


def normalize_column_value(base_col, val):
    """Normaliza un valor de celda para comparar; aplica sinónimos conocidos."""
    s = str(val or "").strip().upper()
    equiv = COLUMN_VALUE_EQUIVALENTS.get(base_col)
    if equiv:
        return equiv.get(s, s)
    return s


def values_equivalent(base_col, old_val, new_val):
    """True si ambos valores representan lo mismo (p. ej. Finalizado = Terminado)."""
    return normalize_column_value(base_col, old_val) == normalize_column_value(base_col, new_val)


def skip_field_on_existing_update(base_col, base_value):
    """Campos que no deben sobrescribirse en filas ya existentes en la base."""
    if base_col in SKIP_ON_EXISTING_UPDATE:
        return True
    if base_col == "ESTADO DE OBRA":
        return normalize_column_value(
            "ESTADO DE OBRA", str(base_value or "").strip()
        ) == "TERMINADO"
    return False


def resolve_column_write_value(base_col, val):
    """Convierte sinónimos PEM al texto canónico de la base (p. ej. Paralizado proyecto → Paralizado)."""
    s = str(val or "").strip()
    if not s:
        return ""
    key = s.upper()
    equiv = COLUMN_VALUE_EQUIVALENTS.get(base_col)
    if not equiv or key not in equiv:
        return s
    canonical_key = equiv[key]
    display = COLUMN_WRITE_CANONICAL.get(base_col, {}).get(canonical_key)
    return display if display else s


def is_percent_100(val):
    """True si el valor representa 100% (1.0 decimal, '100%', etc.)."""
    if isinstance(val, (int, float)):
        return float(val) >= 1.0 - 1e-6
    dec = parse_percent_to_decimal(val)
    return dec is not None and dec >= 1.0 - 1e-6


def pem_cell_is_blank(val):
    """True si la celda PEM viene vacía (no debe sobrescribir la base)."""
    if val is None:
        return True
    return str(val).strip() == ""


def maybe_append_puser_confirmation(changes, mapped_changes, base_row):
    """
    Fija CONFIRMACION DE PROYECTOS (PUSER) con la fecha de hoy solo cuando
    ESTADO DE OBRA pasa a Terminado (no si ya lo era) y % DE AVANCE es 100%.
    """
    if not any(c["field"] == "ESTADO DE OBRA" for c in changes):
        return

    old_estado = str(base_row.get("ESTADO DE OBRA", "") or "").strip()
    if normalize_column_value("ESTADO DE OBRA", old_estado) == "TERMINADO":
        return

    estado_new = mapped_changes.get("ESTADO DE OBRA", "")
    if normalize_column_value("ESTADO DE OBRA", estado_new) != "TERMINADO":
        return

    avance = effective_avance_decimal(mapped_changes, base_row)
    if not is_percent_100(avance):
        return

    today = date.today().strftime("%d/%m/%Y")
    old = str(base_row.get(PUSER_CONFIRMATION_COLUMN, "") or "").strip()
    if old == today:
        return

    changes.append({
        "field": PUSER_CONFIRMATION_COLUMN,
        "before": old or "—",
        "after": today,
    })
    mapped_changes[PUSER_CONFIRMATION_COLUMN] = today


def copy_cell_style(src, dst):
    """Copia estilos básicos de una celda a otra."""
    from copy import copy
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


# ── PEM Endpoints ─────────────────────────────────────────────────────────────

# Almacenamiento temporal en memoria (por sesión de servidor)
_pem_store = {}


@app.route("/api/pem/upload-base", methods=["POST"])
def pem_upload_base():
    """Recibe la base maestra y devuelve metadata de la hoja POBLACIONES."""
    if "file" not in request.files:
        return jsonify({"error": "No se recibió ningún archivo"}), 400

    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Solo se aceptan archivos .xlsx"}), 400

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        SHEET_NAME = "POBLACIONES 2021-2026"
        if SHEET_NAME not in wb.sheetnames:
            return jsonify({"error": f"No se encontró la hoja '{SHEET_NAME}'"}), 400

        ws = wb[SHEET_NAME]

        # Leer headers de la primera fila no vacía
        headers = []
        header_row_idx = None
        for row in ws.iter_rows():
            non_empty = [c for c in row if c.value is not None]
            if non_empty:
                headers = [normalize_header(c.value) for c in row]
                header_row_idx = row[0].row
                break

        if not headers:
            return jsonify({"error": "La hoja no contiene datos"}), 400

        if "EXPEDIENTE" not in headers:
            return jsonify({"error": "No se encontró la columna 'EXPEDIENTE' en la base maestra"}), 400

        # Contar filas de datos
        total_rows = sum(
            1 for row in ws.iter_rows(min_row=header_row_idx + 1)
            if any(c.value is not None for c in row)
        )

        # Guardar en memoria
        _pem_store["base_content"] = content
        _pem_store["base_filename"] = file.filename
        _pem_store["base_headers"] = headers
        _pem_store["base_header_row"] = header_row_idx

        return jsonify({
            "filename": file.filename,
            "total": total_rows,
            "headers": headers,
            "sheetNames": wb.sheetnames,
        })

    except Exception as e:
        return jsonify({"error": f"Error al procesar el archivo: {str(e)}"}), 500


@app.route("/api/pem/upload-avance", methods=["POST"])
def pem_upload_avance():
    """
    Recibe el avance PEM, reconcilia con la base y devuelve preview de cambios.
    """
    if "base_content" not in _pem_store:
        return jsonify({"error": "Primero sube la base maestra"}), 400

    if "file" not in request.files:
        return jsonify({"error": "No se recibió ningún archivo"}), 400

    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Solo se aceptan archivos .xlsx"}), 400

    try:
        # ── Leer avance PEM ───────────────────────────────────────────────────
        pem_content = file.read()
        wb_pem = openpyxl.load_workbook(io.BytesIO(pem_content), data_only=True)

        SHEET_PEM = "Reporte APC 2026"
        if SHEET_PEM not in wb_pem.sheetnames:
            return jsonify({"error": f"No se encontró la hoja '{SHEET_PEM}' en el avance PEM"}), 400

        ws_pem = wb_pem[SHEET_PEM]

        # Saltar primeras 2 filas — headers en fila 3
        all_rows = list(ws_pem.iter_rows(values_only=False))
        if len(all_rows) < 3:
            return jsonify({"error": "El avance PEM no tiene suficientes filas"}), 400

        header_cells = all_rows[2]  # fila índice 2 = fila 3
        pem_headers_raw = [normalize_header(c.value) for c in header_cells]

        # Detectar columna sin nombre al final (para PARALIZADO)
        # Buscar la última columna que tenga valor no vacío en el header
        last_named_col = max(
            (i for i, h in enumerate(pem_headers_raw) if h),
            default=len(pem_headers_raw) - 1
        )
        # La columna "sin nombre" es la siguiente, si existe
        paralizado_col_idx = last_named_col + 1 if last_named_col + 1 < len(pem_headers_raw) else None

        # Verificar que exista N° EXP. en PEM
        if "N° EXP." not in pem_headers_raw:
            return jsonify({"error": "No se encontró la columna 'N° EXP.' en el avance PEM"}), 400

        exp_pem_idx = pem_headers_raw.index("N° EXP.")

        # ── Leer base maestra ─────────────────────────────────────────────────
        wb_base = openpyxl.load_workbook(io.BytesIO(_pem_store["base_content"]), data_only=True)
        ws_base = wb_base["POBLACIONES 2021-2026"]
        base_headers = _pem_store["base_headers"]
        base_header_row = _pem_store["base_header_row"]
        exp_base_idx = base_headers.index("EXPEDIENTE")

        # Construir índice base: expediente → {col: valor}
        base_index = {}
        base_row_map = {}  # expediente → número de fila en Excel
        for row in ws_base.iter_rows(min_row=base_header_row + 1):
            if all(c.value is None for c in row):
                continue
            exp_val = normalize_header(row[exp_base_idx].value) if row[exp_base_idx].value else ""
            if not exp_val:
                continue
            row_data = {base_headers[i]: ("" if cell.value is None else cell.value)
                        for i, cell in enumerate(row) if i < len(base_headers)}
            base_index[exp_val] = row_data
            base_row_map[exp_val] = row[0].row

        # ── Procesar filas PEM ────────────────────────────────────────────────
        updated = []
        added = []
        skipped_paralizado = 0

        for row_cells in all_rows[3:]:  # desde fila 4 en adelante
            if all(c.value is None for c in row_cells):
                continue

            # Chequear columna PARALIZADO
            if paralizado_col_idx is not None and paralizado_col_idx < len(row_cells):
                paralizado_val = str(row_cells[paralizado_col_idx].value or "").strip().upper()
                if "PARALIZADO" in paralizado_val:
                    skipped_paralizado += 1
                    continue

            pem_row = {pem_headers_raw[i]: (c.value if c.value is not None else "")
                       for i, c in enumerate(row_cells) if i < len(pem_headers_raw) and pem_headers_raw[i]}

            exp_val = normalize_header(str(pem_row.get("N° EXP.", "") or ""))
            if not exp_val:
                continue

            # Construir el dict de campos a actualizar (mapeados a nombre base)
            # Para columnas porcentaje, almacenar el valor como decimal
            mapped_changes = {}
            for pem_col, base_col in PEM_TO_BASE.items():
                if pem_col in pem_row and pem_col != "N° EXP.":
                    raw = pem_row[pem_col]
                    if pem_cell_is_blank(raw) and base_col in KEEP_IF_PEM_BLANK:
                        continue
                    if base_col in PERCENT_COLUMNS:
                        mapped_changes[base_col] = parse_pem_avance_to_decimal(raw)
                    else:
                        raw_str = str(raw).strip() if raw != "" else ""
                        mapped_changes[base_col] = resolve_column_write_value(base_col, raw_str)

            if exp_val in base_index:
                # Detectar qué campos realmente cambian
                changes = []
                for base_col, new_val in mapped_changes.items():
                    raw_old = base_index[exp_val].get(base_col, "")
                    if skip_field_on_existing_update(base_col, raw_old):
                        continue

                    if base_col in PERCENT_COLUMNS:
                        old_dec = parse_percent_to_decimal(raw_old)
                        new_dec = new_val  # ya es decimal
                        if new_dec is None and old_dec is not None:
                            continue
                        if (
                            old_dec is not None
                            and new_dec is not None
                            and new_dec < old_dec - 1e-9
                        ):
                            continue
                        # Normalizar: None y 0.0 son equivalentes (celda vacía con formato %)
                        old_norm = old_dec if old_dec is not None else None
                        new_norm = new_dec if new_dec is not None else None
                        if (old_norm or 0.0) != (new_norm or 0.0):
                            changes.append({
                                "field": base_col,
                                "before": format_percent_display(old_dec),
                                "after": format_percent_display(new_dec),
                            })
                    else:
                        old_val = str(raw_old or "").strip()
                        new_val_str = str(new_val).strip() if new_val is not None else ""
                        if not new_val_str and old_val and base_col in KEEP_IF_PEM_BLANK:
                            continue
                        if not values_equivalent(base_col, old_val, new_val_str):
                            changes.append({
                                "field": base_col,
                                "before": old_val,
                                "after": new_val_str,
                            })

                maybe_append_puser_confirmation(changes, mapped_changes, base_index[exp_val])

                if changes:
                    agrupacion = str(base_index[exp_val].get("AGRUPACIONES", exp_val)).strip()
                    updated.append({
                        "expediente": exp_val,
                        "agrupacion": agrupacion,
                        "changes": changes,
                        "mapped": {
                            k: v for k, v in mapped_changes.items()
                            if not skip_field_on_existing_update(
                                k, base_index[exp_val].get(k, "")
                            )
                        },
                    })
            else:
                # Nueva agrupación
                new_row = {}
                for pem_col, base_col in PEM_TO_BASE.items():
                    new_row[base_col] = str(pem_row.get(pem_col, "") or "").strip()
                agrupacion = new_row.get("AGRUPACIONES", exp_val)
                added.append({
                    "expediente": exp_val,
                    "agrupacion": agrupacion,
                    "fields": new_row,
                })

        # Guardar datos para el endpoint de descarga
        _pem_store["pem_content"] = pem_content
        _pem_store["updated"] = updated
        _pem_store["added"] = added
        _pem_store["base_row_map"] = base_row_map
        _pem_store["pem_filename"] = file.filename

        return jsonify({
            "updated": updated,
            "added": added,
            "skippedParalizado": skipped_paralizado,
            "totalPem": len(updated) + len(added) + skipped_paralizado,
        })

    except Exception as e:
        import traceback
        return jsonify({"error": f"Error al procesar el avance: {str(e)}", "detail": traceback.format_exc()}), 500


@app.route("/api/pem/descargar", methods=["POST"])
def pem_descargar():
    """
    Aplica los cambios a la base maestra y devuelve el Excel actualizado,
    conservando formato y todas las hojas.
    """
    if "base_content" not in _pem_store or "updated" not in _pem_store:
        return jsonify({"error": "No hay datos para procesar"}), 400

    try:
        from copy import copy

        # Cargar base con estilos
        wb = openpyxl.load_workbook(io.BytesIO(_pem_store["base_content"]))
        ws = wb["POBLACIONES 2021-2026"]
        base_headers = _pem_store["base_headers"]
        base_header_row = _pem_store["base_header_row"]
        base_row_map = _pem_store["base_row_map"]
        updated = _pem_store["updated"]
        added = _pem_store["added"]

        # Índice columna: nombre → índice (0-based)
        col_idx_map = {h: i for i, h in enumerate(base_headers)}

        # ── Aplicar actualizaciones ───────────────────────────────────────────
        for item in updated:
            exp = item["expediente"]
            row_num = base_row_map.get(exp)
            if row_num is None:
                continue
            for change in item["changes"]:
                base_col = change["field"]
                if base_col not in col_idx_map:
                    continue
                new_val = item["mapped"].get(base_col)
                col_num = col_idx_map[base_col] + 1  # openpyxl es 1-based
                cell = ws.cell(row=row_num, column=col_num)
                if base_col in PERCENT_COLUMNS:
                    # Escribir como decimal; el formato de celda ya es porcentaje
                    cell.value = new_val  # None si vacío, float si tiene valor
                else:
                    cell.value = new_val if new_val != "" else None

        # ── Agregar nuevas filas ──────────────────────────────────────────────
        # Encontrar la última fila con datos
        last_data_row = base_header_row
        for row in ws.iter_rows(min_row=base_header_row + 1):
            if any(c.value is not None for c in row):
                last_data_row = row[0].row

        # Tomar la última fila de datos como plantilla de estilo
        template_row = last_data_row

        for item in added:
            new_row_num = last_data_row + 1
            last_data_row = new_row_num

            for col_idx, base_col in enumerate(base_headers):
                col_num = col_idx + 1
                raw_val = item["fields"].get(base_col, "")
                cell = ws.cell(row=new_row_num, column=col_num)

                # Copiar estilo de la plantilla primero (incluye number_format)
                template_cell = ws.cell(row=template_row, column=col_num)
                copy_cell_style(template_cell, cell)

                if base_col in PERCENT_COLUMNS:
                    cell.value = parse_pem_avance_to_decimal(raw_val)
                else:
                    cell.value = raw_val if raw_val != "" else None

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        base_name = _pem_store.get("base_filename", "base_actualizada.xlsx")
        download_name = base_name.replace(".xlsx", "_actualizada.xlsx")

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )

    except Exception as e:
        import traceback
        return jsonify({"error": f"Error al generar el archivo: {str(e)}", "detail": traceback.format_exc()}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5000)